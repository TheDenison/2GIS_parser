import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from utils.metro_stations import *


def table_exel(data, chat_id):
    # Создание новой книги и листа
    wb = Workbook()
    ws = wb.active

    # Задание шаблона для заголовка
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_alignment = Alignment(horizontal='center', vertical='center')  # добавляем вертикальное выравнивание

    headers = ["ID", "Категория", "Название", "Филиалы", "Адрес", "Сайт", "ИП/ИНН", "Ссылка", "Рейтинг",
               "Оценок", "Телефоны",
               "Email"]
    ws.freeze_panes = 'A2'  # Закрепляет строку 1

    for col_num, header in enumerate(headers, 1):
        col_letter = chr(64 + col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_data)

    # Выравнивание по центру
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center')

    # Автоматическая настройка ширины столбцов
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # # Танируем ячейки
        # for row in ws['G']:
        #     for branch, fill in branches.items():
        #         if any(station in row.value for station in branch):
        #             row.fill = fill
        #             break
    # try:
    #     # Проверка окраски оценки
    #     def interpolate_color(color1, color2, fraction):
    #         r1, g1, b1 = color1
    #         r2, g2, b2 = color2
    #
    #         new_r = int(r1 + (r2 - r1) * fraction)
    #         new_g = int(g1 + (g2 - g1) * fraction)
    #         new_b = int(b1 + (b2 - b1) * fraction)
    #
    #         return new_r, new_g, new_b
    #
    #     def get_color_for_value(value):
    #         RED = (255, 0, 0)
    #         ORANGE = (255, 165, 0)
    #         YELLOW = (255, 255, 0)
    #         LIGHT_GREEN = (124, 252, 0)
    #         GREEN = (0, 255, 0)
    #
    #         value = float(value)
    #
    #         if value <= 1.5:
    #             return RED
    #         elif value <= 2:
    #             return interpolate_color(RED, ORANGE, value - 1.5)
    #         elif value <= 2.5:
    #             return ORANGE
    #         elif value < 3:
    #             return interpolate_color(ORANGE, YELLOW, value - 2.5)
    #         elif value <= 3.5:
    #             return YELLOW
    #         elif value <= 4:
    #             return interpolate_color(YELLOW, LIGHT_GREEN, value - 3.5)
    #         elif value <= 4.5:
    #             return LIGHT_GREEN
    #         elif value < 5:
    #             return interpolate_color(LIGHT_GREEN, GREEN, value - 4.5)
    #         else:
    #             return GREEN
    #
    #     def is_empty(cell_value):
    #         return cell_value is None or str(cell_value).strip() == ""
    #
    #     # Пропустк заголовка и проход по столбцу "J"
    #     for cell in ws['J'][1:]:
    #         if not is_empty(cell.value):
    #             r, g, b = get_color_for_value(cell.value)
    #             fill_color = f"{r:02X}{g:02X}{b:02X}"
    #             cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    # except Exception as e:
    #     print(e)

    # Ссылки для 2ГИС
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):  # столбец H = 7
        cell = row[0]
        url = cell.value
        if url and isinstance(url, str) and url.startswith(('http://', 'https://')):
            cell.hyperlink = url
            cell.style = "Hyperlink"

    # Гиперссылки для сайтов
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):  # Столбец D
        cell = row[0]
        val = str(cell.value).strip()
        if val and '.' in val:
            if not val.startswith(('http://', 'https://')):
                val = 'https://' + val
            cell.hyperlink = val
            cell.value = cell.value  # сохранить отображение оригинального текста
            cell.style = 'Hyperlink'

    cur_time = datetime.datetime.now().strftime('%d_%m_%Y_%H_%M')
    # Сохраняем файл
    wb.save(f"Excels/Table_{cur_time}.xlsx")
    if chat_id != "1":
        wb.save(f"Excels/Table_{chat_id}.xlsx")
