from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def table_exel(data):
    # Создание новой книги и листа
    wb = Workbook()
    ws = wb.active

    # Задание шаблона для заголовка
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_alignment = Alignment(horizontal='center')
    headers = ["ID", "Категория", "Название", "Филиалы", "Адрес", "Сайт", "ИП/ИНН", "Станция", "Ссылка", "Рейтинг",
               "Оценок", "Телефоны",
               "Email"]

    for col_num, header in enumerate(headers, 1):
        col_letter = chr(64 + col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_data)

    # Выравнивание по центру
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center')

    # Автоматическая настройка ширины столбцов
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Сохраняем файл
    wb.save("Table.xlsx")
