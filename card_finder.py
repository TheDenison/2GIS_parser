import datetime
import json
import os
import random
import re
from urllib.parse import urlparse
import pandas as pd
from soup_content import SoupContent
import math
import requests
from bs4 import BeautifulSoup
from time import sleep, time
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from urllib3.util import url
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from constants import districts, type_org_mapping


def filter_links(hrefs):
    pattern = re.compile(r'^https://2gis\.ru/.+/firm/.+')
    return [href for href in hrefs if pattern.match(href)]


# def check_new_firm():
#     with open('firms_dict.json') as file:
#         firms_list = json.load(file)
#     print(firms_list)


class Parser:
    def __init__(self, driver):
        self.driver = driver
        self.soup_data = SoupContent()

    def host(self, urls):
        self.driver.maximize_window()
        self.driver.get('https://2gis.ru/moscow')
        parent_handle = self.driver.window_handles[0]
        org_id = 0
        outputs = []
        for firm_url in urls:
            try:

                self.driver.execute_script(f'window.open("{firm_url}","org_tab");')
                child_handle = [x for x in self.driver.window_handles if x != parent_handle][0]
                self.driver.switch_to.window(child_handle)
                sleep(random.uniform(0.2, 0.8))
                soup = BeautifulSoup(self.driver.page_source, "lxml")

                org_id += 1
                name = self.soup_data.get_name(soup)
                print(f"название: {name}")
                address = self.soup_data.get_address(soup)
                print(f'адрес: {address}')
                link = website = self.soup_data.get_website(soup)
                print(f'сайт: {website}')
                twogis = self.driver.current_url
                print(f'сслыка: {twogis}')
                rating = self.soup_data.get_rating(soup)
                print(f'Оценка: {rating}')
                email = self.soup_data.get_email(soup)
                print(f'Емаил: {email}')
                branch = self.soup_data.get_branch(soup)
                print(f'Филиалы: {branch}')
                ip_inn = ''
                reviews = self.soup_data.get_reviews(soup)
                print(f'Отзывы: {reviews}')
                category = self.soup_data.get_category(soup)
                print(f'Категория: {category}')
                phones = self.get_phones()
                # phone = self.soup_data.get_phone(soup)
                print(f'Номер: {phones}')

                station = self.soup_data.get_station(soup)
                print(f'Станция: {station}')

                if (len(link)) != 0:
                    parsed_url = urlparse(link)
                    if not parsed_url.scheme:
                        link = 'https://' + link
                    ip_inn = self.soup_data.get_ip_inn(link)
                    if ip_inn == None:
                        ip_inn = ''
                    print(f'ИН/ИНН: {ip_inn}\n')

                output = [org_id, category, name, address, website, ip_inn, branch, twogis,
                          rating, reviews, phones, email]
                outputs.append(output)
                if len(outputs) % 50 == 0:
                    self.driver.quit()
                    sleep(random.uniform(2.2, 2.4))
                    self.driver = webdriver.Chrome()
                    self.driver.maximize_window()
                    self.driver.get('https://2gis.ru/moscow')
                    parent_handle = self.driver.window_handles[0]

                self.driver.switch_to.window(parent_handle)
                sleep(random.uniform(0.2, 0.4))

            except:
                print('except')
                self.driver.quit()
                sleep(random.uniform(2.2, 2.4))
                self.driver = webdriver.Chrome()
                self.driver.maximize_window()
                self.driver.get('https://2gis.ru/moscow')
                parent_handle = self.driver.window_handles[0]
        self.table_exel(outputs)
        print('Данные сохранены')
        # self.driver.quit()

    def get_phones(self):
        phones = ""
        try:
            # while not phones:
            self.driver.find_element(By.CLASS_NAME, '_1ns0i7c').click()
            phone_elements = self.driver.find_elements(By.XPATH, '//div[@class="_b0ke8"]/a[contains(@href, "tel:")]')
            phones = [elem.get_attribute('href').replace('tel:', '') for elem in phone_elements]
            phones = '\n'.join(phones)
            sleep(1)
            return phones
        except:
            return phones

    def table_exel(self, data):
        # Создание новой книги и листа
        wb = Workbook()
        ws = wb.active

        # Задание шаблона для заголовка
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")

        headers = ["ID", "Название", "Адрес", "Сайт", "ИП/ИНН", "Остановка", "Ссылка", "Рейтинг", "Отзывы", "Телефон",
                   "Email"]

        for col_num, header in enumerate(headers, 1):
            col_letter = chr(64 + col_num)
            cell = ws[f"{col_letter}1"]
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill

        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_data in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_data)

        # Сохраняем файл
        wb.save("sample.xlsx")
        # Добавляем данные (это можно сделать согласно вашему шаблону)


class CardSearcher:
    def __init__(self, driver, link='https://2gis.ru/moscow'):
        self.driver = driver
        self.slider = None
        self.link = link

    # def make_dirs(self):
    #     directory = f'links/test/{type_org}'
    #     company_collection = 'links/href_collection'
    #     # Проверяем, существует ли файл и директории
    #     if not os.path.exists(company_collection):
    #         os.makedirs(company_collection)

    def check_new_firm(self, city, district, type_org_ru, type_org):
        self.driver.get(self.link)
        self.driver.maximize_window()
        try:
            with open('firms_dict.json') as file:
                firm_list = json.load(file)
        except Exception:
            with open('firms_dict.json', 'w') as file:
                json.dump({}, file)
                firm_list = {}

        fresh_firms = {}
        hrefs_district = []
        # Запрос
        request = city + ' ' + district + ' ' + type_org_ru
        sleep(1)
        find_button = self.driver.find_element(By.CLASS_NAME, '_1gvu1zk')
        find_button.send_keys(request, Keys.ENTER)
        sleep(0.5)
        # включение сортировки по новизне
        find_sort = self.driver.find_element(By.CSS_SELECTOR, 'input._1e4yjns[type="text"][placeholder="Сортировка"]')
        find_sort.send_keys('По новизне', Keys.ENTER)

        sleep(0.3)
        # поиск количества точек
        find_places = self.driver.find_element(By.XPATH, '//div[@class="_nude0k3"]/a[@class="_rdxuhv3"]').text
        number_places = int(''.join(filter(str.isdigit, find_places)))
        print(f'Места: {number_places}')

        # поиск и сбор ссылок на компании
        while '/filters/sort%3Dopened_time' not in self.driver.current_url:
            sleep(0.001)

        find_cards = self.driver.find_element(By.CLASS_NAME, '_z72pvu').find_element(By.CLASS_NAME, '_awwm2v')
        link_elements = find_cards.find_elements(By.CSS_SELECTOR, 'a')
        hrefs = [href.get_attribute('href') for href in link_elements]

        # чистка ссылок от рекламы
        hrefs = filter_links(hrefs)
        hrefs_district += hrefs
        hrefs_district = list(set(hrefs_district))
        for href in hrefs_district:
            clean_url = href.split('?stat')[0]
            firm_id = href.split('/firm/')[1].split('?')[0]
            if firm_id in firm_list:
                continue
            else:
                firm_list[firm_id] = {
                    'url': clean_url
                }

                fresh_firms[firm_id] = {
                    'url': clean_url
                }
        with open("firms_dict.json", 'w') as file:
            json.dump(firm_list, file, indent=1, ensure_ascii=False)

        return fresh_firms


# def first_run():
#     for type_org in ['sport']:
#         i = 0
#         # for district in districts:
#         for district in ['Район сокол']:
#             i += 1
#             print(f"[INFO] Район: {i}/{len(districts)}")
#             driver = webdriver.Chrome()
#             grabber = CardSearcher(driver)
#             grabber.page_parsing(city="Москва", district=district, type_org_ru=type_org_mapping[type_org],
#                                  type_org=type_org)


def update_list():
    fresh_firms_list = {}
    for type_org in ['sport', 'flowers']:
        i = 0

        # for district in districts:
        for district in ['Район сокол']:
            i += 1
            print(f"[INFO] Район: {i}/{len(districts)}")
            driver = webdriver.Chrome()
            grabber = CardSearcher(driver)
            fresh_firms = grabber.check_new_firm(city="Москва", district=district,
                                                 type_org_ru=type_org_mapping[type_org],
                                                 type_org=type_org)
            fresh_firms_list.update(fresh_firms)
            if fresh_firms_list != 0:
                print(f'[Новыe записи] {district}: {len(fresh_firms_list)}')
            # if new_firms != 0:
            #     print(f'Новых записей: {len(new_firms)}')
            #     parse_info()
            #     ss = [i['url'] for i in new_firms.values()]
            #     for s in ss:
            #         print(s)

    with open(f"fresh_firms_dict.json", 'w') as file:
        json.dump(fresh_firms_list, file, indent=1, ensure_ascii=False)
    with open(f"backups/fresh_firms_dict_{cur_time}.json", 'w') as file:
        json.dump(fresh_firms_list, file, indent=1, ensure_ascii=False)


def parse_info():
    try:
        with open('fresh_firms_dict.json') as file:
            firm_list = json.load(file)
    except Exception:
        print('Не получилось открыть файл')
    urls = []
    driver = webdriver.Chrome()
    parser = Parser(driver)
    for k, v in firm_list.items():
        urls.append(f"{v['url']}")
    parser.host(urls)


if __name__ == '__main__':
    start_time = time()
    cur_time = datetime.datetime.now().strftime('%d_%m_%Y_%H_%M')
    # first_run()
    # update_list()
    parse_info()

    end_time = time()
    print(f'Прошло времени: {(end_time - start_time):.1f}')
